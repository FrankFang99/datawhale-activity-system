#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
跑 11 条脱敏数据过 5 维评分引擎，输出分布 + badcase 报告
- 字段映射：v3 修订前（旧字段）→ v3 修订后（新评分输入）
- 输出：
  - 控制台：每条数据的分数 + 等级
  - JSON 报告：分布 + badcase 列表
"""
import json
import subprocess
import sys
import time
from pathlib import Path

# 数据
XLSX = Path(r"D:\Learning\AI\Datawhale\data\test\AI+X创造节_测试数据.xlsx")
REPORT = Path(__file__).parent.parent / "data" / "feishu" / "score_validation_report.json"
REPORT.parent.mkdir(parents=True, exist_ok=True)

# 旧字段 → v3 字段映射
VENUE_MAP = {
    "已确定场地": "已确定",
    "已确定": "已确定",
    "有潜在场地": "有潜在",
    "有潜在": "有潜在",
    "暂时没有，需要支持": "暂无",
    "暂无": "暂无",
    "需要支持": "暂无",
    "": "暂无",
    None: "暂无",
}

# recruitChannel 旧版是"有/无"单选 → v3 多选
def map_recruit_channel(value):
    """旧版单选 → v3 多选"""
    if value in ("有", "yes", "Yes", True):
        return ["社群", "公众号", "高校社团"]  # 假设有 3 个
    if value in ("无", "no", "No", False, None):
        return ["暂无"]
    return ["暂无"]


def read_xlsx():
    """用 Node + xlsx 读 xlsx（避免 Python 装包）"""
    node_script = """
const fs = require('fs');
const path = require('path');
// 简化：手动解析
"""
    # 简单方案：直接调 lark-cli 上传 + 读？太复杂
    # 用 Python 装 openpyxl 已经在用了
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["报名数据表"]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def to_score_input(row):
    """旧字段 → v3 ScoreInput"""
    venue_raw = row.get("是否有预备场地？")
    venue = VENUE_MAP.get(venue_raw, "暂无")
    recruit_raw = row.get("是否有本地招募渠道？")
    recruit = map_recruit_channel(recruit_raw)
    experience = (row.get("介绍你组织过的活动经历（如果是初次见面）") or "").strip()
    motivation = (row.get("你为什么想参与 AI+X 创造节共创？") or "").strip()

    # 预备落地时间 → 解析为 ms timestamp
    expected_date_ms = None
    time_str = row.get("预备落地的时间") or ""
    import re
    # 优先匹配 X月X日
    m = re.search(r"(\d+)\s*月\s*(\d+)\s*日", time_str)
    if m:
        try:
            month, day = int(m.group(1)), int(m.group(2))
            expected_date_ms = int(time.mktime(time.strptime(f"2026-{month:02d}-{day:02d}", "%Y-%m-%d")) * 1000)
        except Exception:
            pass
    # 兜底：8-X（范围形式）
    if expected_date_ms is None:
        m2 = re.search(r"8\s*月\s*(\d+)", time_str)
        if m2:
            try:
                day = int(m2.group(1))
                expected_date_ms = int(time.mktime(time.strptime(f"2026-08-{day:02d}", "%Y-%m-%d")) * 1000)
            except Exception:
                pass
    if expected_date_ms is None:
        # 最终兜底：活动周期 8-2 ~ 8-25，取中间
        expected_date_ms = int(time.mktime(time.strptime("2026-08-15", "%Y-%m-%d")) * 1000)

    # 活动周期（从 活动信息表）
    activity_start = int(time.mktime(time.strptime("2026-08-02", "%Y-%m-%d")) * 1000)
    activity_end = int(time.mktime(time.strptime("2026-08-25", "%Y-%m-%d")) * 1000)

    return {
        "venueStatus": venue,
        "recruitChannel": recruit,
        "experience": experience,
        "expectedDate": expected_date_ms,
        "activityStartDate": activity_start,
        "activityEndDate": activity_end,
        "motivation": motivation,
        "participantValue": "",  # 旧版无此字段
        "_meta": {
            "raw_venue": venue_raw,
            "raw_recruit": recruit_raw,
            "raw_time": time_str,
            "missing_participantValue": True,
        },
    }


def run_score(score_input, retries=3):
    """通过 Node 调 5 维评分引擎（带重试）"""
    last_err = None
    for attempt in range(retries):
        tmp = Path(__file__).parent / "_tmp_score_input.json"
        tmp.write_text(json.dumps(score_input, ensure_ascii=False), encoding="utf-8")
        ts_file = Path(__file__).parent / "_run_score.ts"
        ts_file.write_text(
            "import { scoreApplication } from '../src/modules/score/engine';\n"
            "import * as fs from 'fs';\n"
            "const input = JSON.parse(fs.readFileSync('_tmp_score_input.json', 'utf8'));\n"
            "const result = scoreApplication(input);\n"
            "console.log(JSON.stringify(result));\n",
            encoding="utf-8",
        )
        try:
            npx = "npx.cmd" if sys.platform == "win32" else "npx"
            proc = subprocess.run(
                [npx, "tsx", "_run_score.ts"],
                cwd=str(Path(__file__).parent),
                capture_output=True,
                text=True,
                encoding="utf-8",
                timeout=30,
                shell=True,
            )
            if proc.returncode != 0:
                last_err = proc.stderr[:500]
                continue
            out = proc.stdout.strip()
            for line in out.split("\n")[::-1]:
                line = line.strip()
                if line.startswith("{"):
                    try:
                        return json.loads(line), None
                    except Exception as e:
                        last_err = f"json parse: {e}"
                        continue
            last_err = out[-500:]
        finally:
            if ts_file.exists():
                ts_file.unlink()
            if tmp.exists():
                tmp.unlink()
        time.sleep(0.5)  # 短暂延迟避免 spawn 太快
    return None, last_err or "unknown error after retries"


def main():
    print("📊 11 条脱敏数据 5 维评分验证\n")
    rows = read_xlsx()
    print(f"读取 {len(rows)} 条数据\n")

    results = []
    distribution = {"S": 0, "A": 0, "B": 0, "C": 0, "D": 0}
    badcases = []
    score_sum = 0

    for i, row in enumerate(rows, 1):
        score_input = to_score_input(row)
        result, err = run_score(score_input)
        if err or not result:
            print(f"  [{i}] {row.get('姓名')}: ❌ err={err or 'None result'}", flush=True)
            badcases.append({"row": i, "name": row.get("姓名"), "error": (err or "None result")[:200]})
            continue
        try:
            score = result["total"]
            grade = result["grade"]
            rc1 = result["RC001"]["score"]
            rc2 = result["RC002"]["score"]
            rc3 = result["RC003"]["score"]
            rc4 = result["RC004"]["score"]
            rc5 = result["RC005"]["score"]
        except Exception as e:
            print(f"  [{i}] {row.get('姓名')}: ❌ missing fields: {e}, result={result}", flush=True)
            badcases.append({"row": i, "name": row.get("姓名"), "error": f"missing fields: {e}"})
            continue
        distribution[grade] += 1
        score_sum += score
        results.append({
            "row": i,
            "name": row.get("姓名"),
            "school": row.get("您的学校"),
            "city": row.get("您的现居城市"),
            "raw_venue": score_input["_meta"]["raw_venue"],
            "raw_recruit": score_input["_meta"]["raw_recruit"],
            "score": score,
            "grade": grade,
            "RC001": rc1,
            "RC002": rc2,
            "RC003": rc3,
            "RC004": rc4,
            "RC005": rc5,
        })
        # badcase 标记
        issues = []
        if not score_input["_meta"]["raw_recruit"]:
            issues.append("recruitChannel 字段缺失（旧版问卷）")
        if not score_input["_meta"]["raw_time"]:
            issues.append("expectedDate 解析失败")
        if score_input["_meta"]["missing_participantValue"]:
            issues.append("participantValue 字段缺失（v3 新增）")
        if result["RC001"]["score"] == 0 and result["RC001"]["reason"].endswith("异常"):
            issues.append(f"RC-001 非法值：{score_input['_meta']['raw_venue']}")

        status = "🟢" if not issues else "🟡"
        name = row.get("姓名") or "?"
        school = row.get("您的学校") or "-"
        print(f"  [{i:2}] {status} {name[:6]:8s} | {school[:10]:10s} | {grade} {score:3d} | RC: {rc1}/{rc2}/{rc3}/{rc4}/{rc5}", flush=True)
        if issues:
            for issue in issues:
                print(f"        ⚠️  {issue}", flush=True)
            badcases.append({"row": i, "name": row.get("姓名"), "issues": issues, "score": score, "grade": grade})

    print(f"\n📈 分布 (n={len(results)}):")
    for g, c in distribution.items():
        pct = c / len(results) * 100 if results else 0
        bar = "█" * c
        print(f"  {g}: {c:2d} ({pct:5.1f}%) {bar}")
    if results:
        print(f"\n平均分: {score_sum / len(results):.1f}")

    print(f"\n⚠️  Badcase 报告: {len(badcases)} 条")
    for bc in badcases:
        print(f"  - [{bc['row']}] {bc['name']}: {bc.get('issues', bc.get('error', '?'))[:200]}")

    # 保存 JSON 报告
    report = {
        "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "totalRows": len(rows),
        "scoredRows": len(results),
        "distribution": distribution,
        "averageScore": score_sum / len(results) if results else 0,
        "results": results,
        "badcases": badcases,
        "notes": [
            "11 条脱敏数据来自 5-30 ~ 7-12 收集，是 §4.1.4 v3 改稿前版本",
            "字段映射：'是否有预备场地？' → venueStatus，'是否有本地招募渠道？' → recruitChannel",
            "recruitChannel 旧版单选'有/无' → v3 多选'有=3渠道/无=暂无'",
            "participantValue 字段缺失（旧版问卷无此字段）",
            "expectedDate 文本解析为 2026-M-D（兜底 8-15）",
        ],
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n✅ 报告已保存: {REPORT}")


if __name__ == "__main__":
    main()
